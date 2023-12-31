from random import choice, randint
from openpyxl import Workbook, load_workbook
#生成随机数据
def generateRandomInformation(filename):
    workbook = Workbook()
    worksheet = workbook.worksheets[0]
    worksheet.append(['姓名','课程','成绩'])
    #中文名字中的第一、第二、第三个字
    first = '赵钱孙李'
    middle = '伟昀琛东'
    last = '坤艳志'
    subjects = ('语文','数学','英语')
    for i in range(200):
        name = choice(first)
        #按一定概率生成只有两个字的中文名字
        if randint(1,100)>50:
            name = name + choice(middle)
        name = name + choice(last)
        #依次生成姓名、课程名称和成绩
        worksheet.append([name, choice(subjects), randint(0, 100)])
    #保存数据，生成Excel 2007 格式的文件
    workbook.save(filename)
    
def getResult(oldfile, newfile):
    #用于存放结果数据的字典
    result = dict()
    #打开原始数据
    workbook = load_workbook(oldfile)
    worksheet = workbook.worksheets[0]
    #遍历原始数据
    for row in worksheet.rows:
        if row[0].value == '姓名':
            continue
        #姓名,课程名称,本次成绩
        name, subject, grade = map(lambda cell:cell.value, row)
        #获取当前姓名对应的课程名称和成绩信息
        #如果result 字典中不包含，则返回空字典
        t = result.get(name, {})
        #获取当前学生当前课程的成绩，若不存在，返回0
        f = t.get(subject, 0)
        #只保留该学生该课程的最高成绩
        if grade > f:
            t[subject] = grade
            result[name] = t
    workbook1 = Workbook()
    worksheet1 = workbook1.worksheets[0]
    worksheet1.append(['姓名','课程','成绩'])
    #将result 字典中的结果数据写入Excel 文件
    for name, t in result.items():
        print(name, t)
        for subject, grade in t.items():
            worksheet1.append([name, subject, grade])
    workbook1.save(newfile)
    

oldfile = r'test.xlsx'
newfile = r'result.xlsx'

generateRandomInformation(oldfile)
getResult(oldfile, newfile)


#以下为实验内容
#一、获取语文成绩排名
wb = load_workbook('result.xlsx')
ws = wb[wb.sheetnames[0]] 
scores = {}
for row in ws.iter_rows(min_row=2, values_only=True):
    name, course, score = row
    if course == "语文":
        scores[name] = score
sorted_scores = sorted(scores.items(), key=lambda x: x[1], reverse=True)

print("语文课程所有学生的成绩（按成绩从高到低排序）：")
for name, score in sorted_scores:
    print(f"{name}: {score}")

#二、统计数学不及格的学生的名单
failed_students = []
for row in ws.iter_rows(min_row=2, values_only=True):
    name, course, score = row
    if course == "数学" and score < 60:
        failed_students.append(name)

print("数学不及格的学生名单：")
print(failed_students)

#三、统计总分前五的学生
import heapq
total_scores = {}
for row in ws.iter_rows(min_row=2, values_only=True):
    name, course, score = row
    if name not in total_scores:
        total_scores[name] = 0
    total_scores[name] += score

top_5 = heapq.nlargest(5, total_scores.items(), key=lambda x: x[1])

print("总分最高的前五名：")
for i, (name, score) in enumerate(top_5):
    print(f"{i+1}. {name}: {score}")